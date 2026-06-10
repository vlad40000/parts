"""
Appliance parts-list pipeline (washers, dryers, fridges, ranges, freezers).

Pure Gemini flow: no scraping libraries (no BeautifulSoup / cheerio / playwright).
Data acquisition is done entirely through Gemini's built-in google_search and
url_context tools. Python here is orchestration glue only: it spawns the parallel
model calls, computes exact-string dedup, and runs the loop.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sys
import uuid
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen

from google import genai
from google.genai import types

try:
    import openpyxl  # optional; only required for .xlsx export
except ImportError:
    openpyxl = None


# ----------------------------------------------------------------------------- config
MODEL_FLASH = "gemini-3.5-flash"
MODEL_LITE = "gemini-3.1-flash-lite"
MODEL_PRO = "gemini-3.1-pro-preview"

# Loop safety. Callout counts over/under-estimate, so never trust the target alone.
MAX_ROUNDS = 8          # hard ceiling on step 6/7 iterations
STALL_ROUNDS = 2        # stop if this many consecutive rounds add zero new parts
WORKERS_STEP2 = 3
WORKERS_STEP4 = 3
WORKERS_STEP6 = 5
PARTS_PER_WORKER = 50
MAX_CREDIBLE_EXPECTED_PARTS = 500
TRUSTED_EXPECTED_COUNT_HOSTS = frozenset({
    "encompass.com",
    "www.encompass.com",
    "searspartsdirect.com",
    "www.searspartsdirect.com",
})
SOURCE_FETCH_TIMEOUT_SECONDS = 10
SOURCE_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/137.0.0.0 Safari/537.36"
)

client = None


# ---------------------------------------------------------------------------- schemas
S1_NAMEPLATE = {
    "type": "object",
    "required": ["model_number", "serial_number", "extraction_confidence"],
    "properties": {
        "model_number": {"type": "string"},
        "serial_number": {"type": ["string", "null"]},
        "brand": {"type": ["string", "null"]},
        "appliance_type": {"type": ["string", "null"]},
        "extraction_confidence": {"type": "string", "enum": ["high", "medium", "low"]},
        "notes": {"type": ["string", "null"]},
    },
}

S2_DIAGRAMS = {
    "type": "object",
    "required": ["diagrams"],
    "properties": {
        "diagrams": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["section_name", "page_url", "source_site"],
                "properties": {
                    "section_name": {"type": "string"},
                    "diagram_url": {"type": ["string", "null"]},   # direct image if resolvable
                    "page_url": {"type": "string"},
                    "source_site": {"type": "string"},
                    "distinct_callout_count_seen": {"type": ["integer", "null"]},
                    "max_reference_label_seen": {"type": ["string", "null"]},
                    "notes": {"type": ["string", "null"]},
                },
            },
        }
    },
}

S3_SECTIONS = {
    "type": "object",
    "required": ["sections"],
    "properties": {
        "sections": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["section_name"],
                "properties": {
                    "section_name": {"type": "string"},
                    "aliases": {"type": "array", "items": {"type": "string"}},
                    "diagram_urls": {"type": "array", "items": {"type": "string"}},
                    "page_urls": {"type": "array", "items": {"type": "string"}},
                    "observed_part_count": {"type": ["integer", "null"]},
                    "max_reference_label": {"type": ["string", "null"]},
                },
            },
        }
    },
}

S4_AUDIT = {
    "type": "object",
    "required": ["missing_sections"],
    "properties": {
        "missing_sections": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["section_name"],
                "properties": {
                    "section_name": {"type": "string"},
                    "reason": {"type": ["string", "null"]},
                    "diagram_url": {"type": ["string", "null"]},
                    "page_url": {"type": ["string", "null"]},
                },
            },
        }
    },
}

S5_COUNT = {
    "type": "object",
    "required": ["expected_parts_count", "confidence", "basis"],
    "properties": {
        "expected_parts_count": {"type": "integer"},
        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
        "basis": {
            "type": "string",
            "enum": ["diagram_callouts", "source_total", "cross_referenced"],
        },
        "source_totals": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["source", "count", "url"],
                "properties": {
                    "source": {"type": "string"},
                    "count": {"type": "integer"},
                    "url": {"type": "string"},
                    "evidence": {"type": ["string", "null"]},
                },
            },
        },
        "per_section_counts": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["section_name", "count"],
                "properties": {
                    "section_name": {"type": "string"},
                    "count": {"type": "integer"},
                },
            },
        },
        "notes": {"type": ["string", "null"]},
    },
}

# Step 6 worker schema (matches the standalone schema file shared earlier).
S6_WORKER = {
    "type": "object",
    "required": ["worker_id", "model_number", "assigned_scope", "parts_returned", "parts"],
    "properties": {
        "worker_id": {"type": "string"},
        "model_number": {"type": "string"},
        "assigned_scope": {"type": "array", "items": {"type": "string"}},
        "parts_returned": {"type": "integer"},
        "parts": {
            "type": "array",
            "items": {
                "type": "object",
                "required": [
                    "section",
                    "diagram_callout_id",
                    "manufacturer_part_number",
                    "part_number_status",
                    "part_name",
                    "source_url",
                ],
                "properties": {
                    "section": {"type": "string"},
                    "diagram_callout_id": {"type": ["string", "null"]},
                    "manufacturer_part_number": {"type": ["string", "null"]},
                    "part_number_status": {
                        "type": "string",
                        "enum": ["confirmed", "probable", "unconfirmed"],
                    },
                    "part_name": {"type": "string"},
                    "quantity": {"type": ["integer", "null"]},
                    "is_sub_assembly": {"type": "boolean"},
                    "source_url": {"type": "string"},
                    "diagram_url": {"type": ["string", "null"]},
                    "notes": {"type": ["string", "null"]},
                },
            },
        },
    },
}

S7_GAMEPLAN = {
    "type": "object",
    "required": ["assignments", "still_missing_estimate"],
    "properties": {
        "summary": {"type": ["string", "null"]},
        "still_missing_estimate": {"type": "integer"},
        "assignments": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["worker_id", "target_sections"],
                "properties": {
                    "worker_id": {"type": "string"},
                    "target_sections": {"type": "array", "items": {"type": "string"}},
                    "callout_ranges": {"type": "array", "items": {"type": "string"}},
                    "focus_note": {"type": ["string", "null"]},
                    # callout/section refs of unconfirmed parts to nail down next round
                    "resolve_unconfirmed": {"type": "array", "items": {"type": "string"}},
                },
            },
        },
    },
}


# ---------------------------------------------------------------------------- helpers
_TOOLS = [
    types.Tool(url_context=types.UrlContext()),
    types.Tool(google_search=types.GoogleSearch()),
]


def _schema_field(schema: dict) -> dict:
    """Pick whichever response-schema kwarg the installed SDK version exposes."""
    try:
        fields = types.GenerateContentConfig.model_fields
        if "response_json_schema" in fields:
            return {"response_json_schema": schema}
    except Exception:
        pass
    return {"response_schema": schema}


def _loads(text: str) -> dict:
    """Parse model JSON. Structured output shouldn't fence it, but strip just in case."""
    if text is None:
        raise ValueError("empty model response")
    cleaned = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
    return json.loads(cleaned)


def _config(schema: dict, thinking_level: str) -> types.GenerateContentConfig:
    return types.GenerateContentConfig(
        thinking_config=types.ThinkingConfig(thinking_level=thinking_level),
        tools=_TOOLS,
        response_mime_type="application/json",
        **_schema_field(schema),
    )


def _call(model: str, prompt: str, schema: dict, thinking_level: str,
          image_bytes: bytes | None = None, image_mime: str = "image/jpeg",
          retries: int = 1) -> dict:
    """One structured Gemini call. Optionally attaches a local image (nameplate photo)."""
    parts: list = []
    if image_bytes is not None:
        parts.append(types.Part.from_bytes(data=image_bytes, mime_type=image_mime))
    parts.append(types.Part.from_text(text=prompt))
    contents = [types.Content(role="user", parts=parts)]

    last_err = None
    for _ in range(retries + 1):
        try:
            global client
            if client is None:
                api_key = os.environ.get("GEMINI_API_KEY")
                client = genai.Client(api_key=api_key) if api_key else genai.Client()
            resp = client.models.generate_content(
                model=model, contents=contents,
                config=_config(schema, thinking_level),
            )
            return _loads(resp.text)
        except Exception as e:  # transient network / parse hiccup -> one retry
            last_err = e
    raise RuntimeError(f"{model} call failed: {last_err}")


def _parallel(jobs: list) -> list:
    """Run a list of zero-arg callables concurrently; return results in submit order."""
    results: list = [None] * len(jobs)
    with ThreadPoolExecutor(max_workers=len(jobs)) as ex:
        futs = {ex.submit(job): i for i, job in enumerate(jobs)}
        for fut in as_completed(futs):
            i = futs[fut]
            try:
                results[i] = fut.result()
            except Exception as e:
                results[i] = {"_error": str(e)}
    return results


# ------------------------------------------------------------------------------ steps
def step1_nameplate(model_number=None, serial=None, brand=None, appliance_type=None,
                    raw_text=None, image_path=None) -> dict:
    """Manual entry short-circuits the model. Image/text input goes through gemini-3.5-flash."""
    if image_path is None and raw_text is None and model_number:
        return {
            "model_number": model_number, "serial_number": serial,
            "brand": brand, "appliance_type": appliance_type,
            "extraction_confidence": "high", "notes": "manual entry",
        }

    img = open(image_path, "rb").read() if image_path else None
    prompt = (
        "You are reading an appliance nameplate/rating label. Extract the MODEL NUMBER "
        "and SERIAL NUMBER (both required), plus BRAND and APPLIANCE TYPE if visible. "
        "Use google_search/url_context to disambiguate the brand or model format if the "
        "label text is partially legible. Do NOT guess a serial number you cannot read: "
        "return null and set extraction_confidence to 'low'. "
        + (f"\n\nLabel text provided:\n{raw_text}" if raw_text else "")
    )
    return _call(MODEL_FLASH, prompt, S1_NAMEPLATE, "MINIMAL", image_bytes=img)


def step2_find_diagrams(nameplate: dict) -> list:
    """3 parallel diagram searches, lightly diversified by source preference."""
    mn = nameplate["model_number"]
    brand = nameplate.get("brand") or ""
    atype = nameplate.get("appliance_type") or ""
    leans = [
        "Prefer the manufacturer's own parts/service site and OEM diagrams.",
        "Prefer major parts retailers (e.g. searspartsdirect, repairclinic, partselect).",
        "Prefer appliance-repair community diagrams and service manuals.",
    ]

    def make(lean):
        def job():
            prompt = (
                f"Find exploded-view / assembly parts diagrams for appliance model '{mn}' "
                f"{('brand ' + brand) if brand else ''} {('type ' + atype) if atype else ''}. "
                "Use google_search to locate diagram pages, then url_context to OPEN each page "
                "or diagram image and read it. For every distinct assembly section (e.g. "
                "'Drum', 'Control Panel', 'Door', 'Cabinet'), return the section name, the page "
                "url, the direct diagram image url if you can resolve it, the source site, and "
                "the number of DISTINCT visible callout labels. Also preserve the highest visible "
                "reference label separately as max_reference_label_seen. Reference labels such as "
                "999 or 825 are identifiers, not counts. "
                f"{lean} Only report diagrams that actually match model '{mn}' or an explicitly "
                "stated compatible model."
            )
            return _call(MODEL_FLASH, prompt, S2_DIAGRAMS, "MINIMAL")
        return job

    return _parallel([make(l) for l in leans])


def step3_consolidate(nameplate: dict, diagram_results: list) -> dict:
    mn = nameplate["model_number"]
    blob = json.dumps([r for r in diagram_results if "_error" not in r], ensure_ascii=False)
    prompt = (
        f"These are diagram findings from three searches for model '{mn}'. Merge them into one "
        "canonical list of assembly sections. Deduplicate sections that are the same thing under "
        "different names (record the variants in 'aliases'). For each canonical section, union all "
        "diagram_urls and page_urls, keep the best distinct observed_part_count, and preserve the "
        "largest max_reference_label only as an identifier. Return only "
        f"sections that plausibly belong to model '{mn}'.\n\nFindings:\n{blob}"
    )
    return _call(MODEL_LITE, prompt, S3_SECTIONS, "MINIMAL")


def consolidate_diagram_results(diagram_results: list) -> dict:
    """Merge exact diagram findings without spending another model call."""
    merged = {}
    for result in diagram_results:
        if not isinstance(result, dict) or "_error" in result:
            continue
        for diagram in result.get("diagrams", []) or []:
            name = (diagram.get("section_name") or "").strip()
            if not name:
                continue
            key = re.sub(r"[^a-z0-9]+", "", name.lower())
            section = merged.setdefault(key, {
                "section_name": name,
                "aliases": [],
                "diagram_urls": [],
                "page_urls": [],
                "observed_part_count": None,
                "max_reference_label": None,
            })
            if name != section["section_name"] and name not in section["aliases"]:
                section["aliases"].append(name)
            for field, target in (
                ("diagram_url", "diagram_urls"),
                ("page_url", "page_urls"),
            ):
                value = diagram.get(field)
                if value and value not in section[target]:
                    section[target].append(value)
            observed = _credible_count(diagram.get("distinct_callout_count_seen"))
            if observed is not None:
                current = section["observed_part_count"] or 0
                section["observed_part_count"] = max(current, observed)
            reference = diagram.get("max_reference_label_seen")
            if reference:
                section["max_reference_label"] = str(reference)
    return {"sections": list(merged.values())}


def step4_audit(nameplate: dict, sections: dict) -> list:
    """3 parallel audits checking whether any section was missed."""
    mn = nameplate["model_number"]
    have = json.dumps([s["section_name"] for s in sections.get("sections", [])])

    def job():
        prompt = (
            f"For appliance model '{mn}' we currently have these assembly sections: {have}. "
            "Using google_search and url_context against the model's parts catalog, identify any "
            "STANDARD section for this appliance type that is MISSING from the list (e.g. wiring "
            "harness, motor, pump, hinge/closure, trim, optional accessories). For each missing "
            "section give its name, a reason, and a page or diagram url if you find one. If nothing "
            "is missing, return an empty array."
        )
        return _call(MODEL_LITE, prompt, S4_AUDIT, "MINIMAL")

    return _parallel([job for _ in range(WORKERS_STEP4)])


def merge_audit_into_sections(sections: dict, audits: list) -> dict:
    known = {s["section_name"].strip().lower() for s in sections.get("sections", [])}
    for a in audits:
        for m in a.get("missing_sections", []) if "_error" not in a else []:
            name = m.get("section_name", "").strip()
            has_model_evidence = bool(m.get("diagram_url") or m.get("page_url"))
            if name and has_model_evidence and name.lower() not in known:
                known.add(name.lower())
                sections["sections"].append({
                    "section_name": name,
                    "aliases": [],
                    "diagram_urls": [m["diagram_url"]] if m.get("diagram_url") else [],
                    "page_urls": [m["page_url"]] if m.get("page_url") else [],
                    "observed_part_count": None,
                    "max_reference_label": None,
                })
    return sections


def _normalize_model_token(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _fetch_source_html(url: str) -> str:
    request = Request(
        url,
        headers={
            "User-Agent": SOURCE_USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    with urlopen(request, timeout=SOURCE_FETCH_TIMEOUT_SECONDS) as response:
        return response.read().decode("utf-8", errors="replace")


def _extract_balanced_json(text: str, marker: str) -> dict | None:
    marker_index = text.find(marker)
    if marker_index < 0:
        return None
    first_brace = text.find("{", marker_index)
    if first_brace < 0:
        return None

    depth = 0
    in_string = False
    escaped = False
    for index in range(first_brace, len(text)):
        char = text[index]
        if escaped:
            escaped = False
            continue
        if char == "\\" and in_string:
            escaped = True
            continue
        if char == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(text[first_brace:index + 1])
                except json.JSONDecodeError:
                    return None
    return None


def _dynamic_count(record: dict, prefix: str) -> int | None:
    for key, value in record.items():
        if not str(key).startswith(prefix):
            continue
        if isinstance(value, int):
            return _credible_count(value)
        if isinstance(value, dict):
            return _credible_count(value.get("totalCount"))
    return None


def parse_sears_expected_count(html: str, model_number: str) -> dict | None:
    payload = _extract_balanced_json(html, "CATALOG_API_RESPONSE")
    if not payload:
        return None

    requested = _normalize_model_token(model_number)
    candidates = []
    for value in payload.values():
        if not isinstance(value, dict) or value.get("__typename") != "Model":
            continue
        candidate_model = _normalize_model_token(
            value.get("number") or value.get("modelNumber") or value.get("model")
        )
        if candidate_model != requested:
            continue
        count = (
            _dynamic_count(value, "partCount(")
            or _credible_count(value.get("partCount"))
        )
        if count:
            candidates.append(count)

    if not candidates:
        return None
    count = max(candidates)
    url = f"https://www.searspartsdirect.com/search?q={quote(model_number)}"
    return {
        "source": "Sears PartsDirect",
        "count": count,
        "url": url,
        "evidence": f"Exact model {model_number}: {count} parts",
    }


def parse_encompass_expected_count(html: str, model_number: str) -> dict | None:
    requested = _normalize_model_token(model_number)
    compact_html = re.sub(r"\s+", " ", html)
    model_matches = list(re.finditer(re.escape(model_number), compact_html, re.IGNORECASE))
    if not model_matches or requested not in _normalize_model_token(compact_html):
        return None

    counts = []
    for model_match in model_matches:
        end = min(len(compact_html), model_match.end() + 1200)
        nearby = compact_html[model_match.end():end]
        count_match = re.search(r"\b(\d{1,3})\s+parts?\b", nearby, re.IGNORECASE)
        if count_match and (count := _credible_count(count_match.group(1))) is not None:
            counts.append(count)
    if not counts:
        return None

    count = max(counts)
    url = f"https://encompass.com/search?searchTerm={quote(model_number)}"
    return {
        "source": "Encompass",
        "count": count,
        "url": url,
        "evidence": f"Exact model {model_number}: {count} parts",
    }


def _direct_expected_count_source(
    source_name: str,
    url: str,
    parser,
    model_number: str,
) -> dict:
    try:
        html = _fetch_source_html(url)
        result = parser(html, model_number)
        return result or {
            "_error": f"{source_name} did not expose an exact-model part count"
        }
    except (HTTPError, URLError, TimeoutError, OSError) as exc:
        return {"_error": f"{source_name} fetch failed: {exc}"}


def _grounded_expected_count_source(
    source_name: str,
    source_url: str,
    model_number: str,
) -> dict:
    prompt = (
        f"Open this exact {source_name} URL for appliance model '{model_number}': {source_url}\n"
        f"Return the visible number labeled as parts for the exact model '{model_number}'. "
        "Do not use diagram callout numbers, search-result totals for other models, compatible "
        "models, or any other domain. source_totals must be empty if the exact model and its "
        "visible parts count cannot be verified from this site. When verified, include exactly "
        "one source_totals item with this URL and evidence text containing the model and count. "
        "Set expected_parts_count to that count, basis=source_total, and confidence=high."
    )
    return _call(MODEL_FLASH, prompt, S5_COUNT, "MINIMAL", retries=1)


def step5_expected_count(nameplate: dict, sections: dict | None = None) -> dict:
    mn = nameplate["model_number"]
    sources = [
        (
            "Sears PartsDirect",
            f"https://www.searspartsdirect.com/search?q={quote(mn)}",
            parse_sears_expected_count,
        ),
        (
            "Encompass",
            f"https://encompass.com/search?searchTerm={quote(mn)}",
            parse_encompass_expected_count,
        ),
    ]
    direct_results = _parallel([
        (
            lambda source_name=source_name, url=url, parser=parser:
            _direct_expected_count_source(source_name, url, parser, mn)
        )
        for source_name, url, parser in sources
    ])
    source_totals = [
        result for result in direct_results
        if isinstance(result, dict) and "_error" not in result
    ]

    grounded_results = []
    if len(source_totals) < len(sources):
        resolved_hosts = {
            (urlparse(source.get("url") or "").hostname or "").lower()
            for source in source_totals
        }
        unresolved_sources = [
            (source_name, url)
            for source_name, url, _parser in sources
            if (urlparse(url).hostname or "").lower() not in resolved_hosts
        ]
        grounded_results = _parallel([
            (
                lambda source_name=source_name, url=url:
                _grounded_expected_count_source(source_name, url, mn)
            )
            for source_name, url in unresolved_sources
        ])
        for result in grounded_results:
            if not isinstance(result, dict) or "_error" in result:
                continue
            for source in result.get("source_totals", []) or []:
                if _credible_count(source.get("count")) and _trusted_expected_count_source(source):
                    source_totals.append(source)

    unique_sources = {}
    for source in source_totals:
        key = (
            (urlparse(source.get("url") or "").hostname or "").lower(),
            _credible_count(source.get("count")),
        )
        unique_sources[key] = source
    source_totals = list(unique_sources.values())

    if not source_totals:
        errors = [
            result.get("_error")
            for result in [*direct_results, *grounded_results]
            if isinstance(result, dict) and result.get("_error")
        ]
        raise RuntimeError(
            "Unable to establish an exact-model expected part count from "
            f"Sears PartsDirect or Encompass for {mn}. "
            + " ".join(errors)
        )

    values = [
        _credible_count(source.get("count"))
        for source in source_totals
        if _credible_count(source.get("count")) is not None
    ]
    expected = max(values)
    return {
        "expected_parts_count": expected,
        "confidence": "high",
        "basis": "cross_referenced" if len(source_totals) > 1 else "source_total",
        "source_totals": source_totals,
        "per_section_counts": [
            {
                "section_name": section.get("section_name"),
                "count": section.get("observed_part_count"),
            }
            for section in (sections or {}).get("sections", [])
            if _credible_count(section.get("observed_part_count"))
        ],
        "notes": (
            f"Exact-model source totals disagree; using {expected} as the coverage target."
            if len(set(values)) > 1
            else "Exact-model source count verified."
        ),
    }


def _credible_count(value) -> int | None:
    try:
        count = int(value)
    except (TypeError, ValueError):
        return None
    return count if 0 < count <= MAX_CREDIBLE_EXPECTED_PARTS else None


def _trusted_expected_count_source(source: dict) -> bool:
    try:
        host = (urlparse(source.get("url") or "").hostname or "").lower()
    except ValueError:
        return False
    return host in TRUSTED_EXPECTED_COUNT_HOSTS


def normalize_expected_count(count_info: dict) -> tuple[int, dict]:
    """
    Select a bounded expected count from exact-model evidence.

    Source totals outrank model-computed totals. Diagram reference labels are
    never considered because they are identifiers, not row counts.
    """
    source_counts = []
    rejected_source_counts = []
    for source in count_info.get("source_totals", []) or []:
        count = _credible_count(source.get("count"))
        normalized = {
            "source": source.get("source"),
            "count": count,
            "url": source.get("url"),
            "evidence": source.get("evidence"),
        }
        if count is not None and _trusted_expected_count_source(source):
            source_counts.append(normalized)
        elif count is not None:
            rejected_source_counts.append(normalized)

    raw_count = count_info.get("expected_parts_count")
    raw_credible = _credible_count(raw_count)
    selected = 0
    selection_basis = "unknown"

    if source_counts:
        selected = max(source["count"] for source in source_counts)
        selection_basis = "source_totals"

    section_counts = [
        count
        for item in count_info.get("per_section_counts", []) or []
        if (count := _credible_count(item.get("count"))) is not None
    ]
    meta = {
        **count_info,
        "raw_expected_parts_count": raw_count,
        "selected_expected_parts_count": selected,
        "selection_basis": selection_basis,
        "source_totals": source_counts,
        "rejected_source_totals": rejected_source_counts,
        "diagram_occurrence_total": sum(section_counts) if section_counts else None,
    }
    if source_counts:
        values = [source["count"] for source in source_counts]
        meta["credible_source_range"] = {
            "minimum": min(values),
            "maximum": max(values),
        }
    if raw_count and (raw_credible is None or not source_counts):
        meta["rejected_expected_parts_count"] = raw_count

    return selected, meta


def step6_extract(nameplate: dict, sections: dict, assignment: dict, found_keys: set) -> dict:
    """One worker extracts up to 20 parts within its assigned lane."""
    mn = nameplate["model_number"]
    sec_index = json.dumps(sections.get("sections", []), ensure_ascii=False)
    avoid = json.dumps(sorted(list(found_keys))[:400])  # cap prompt size
    prompt = (
        f"You are worker {assignment['worker_id']} building the parts list for model '{mn}'. "
        f"Your assigned scope (stay strictly inside it): {json.dumps(assignment.get('target_sections', []))} "
        f"callout ranges {json.dumps(assignment.get('callout_ranges', []))}. "
        f"Find up to {PARTS_PER_WORKER} DISTINCT parts within that scope.\n\n"
        "Method: open the relevant diagram(s) via url_context to see the callouts, then use "
        "google_search / url_context on the model's parts table to map each visible callout id to "
        "its orderable MANUFACTURER PART NUMBER.\n\n"
        "CRITICAL RULES:\n"
        "- NEVER invent a part number. If you cannot locate the real MPN, return it with "
        "manufacturer_part_number=null and part_number_status='unconfirmed'.\n"
        "- part_number_status: 'confirmed' = read from THIS model's parts table; 'probable' = from "
        "an explicitly compatible model; 'unconfirmed' = no MPN found.\n"
        "- Every part needs the source_url where you read it.\n"
        "- If one callout represents a kit/box of multiple parts, set is_sub_assembly=true.\n"
        f"- Do NOT return any part whose MPN or (section+callout) already appears here: {avoid}\n\n"
        "Echo your worker_id, the model_number, your assigned_scope, and parts_returned.\n\n"
        f"Section index:\n{sec_index}"
    )
    return _call(MODEL_LITE, prompt, S6_WORKER, "MINIMAL")


def step7_gameplan(nameplate: dict, sections: dict, master: dict,
                   expected: int, round_no: int, past_covered: list) -> dict:
    mn = nameplate["model_number"]
    by_status = {"confirmed": 0, "probable": 0, "unconfirmed": 0}
    unconfirmed_refs = []
    coverage = {}
    for p in master.values():
        by_status[p.get("part_number_status", "unconfirmed")] += 1
        coverage[p.get("section", "?")] = coverage.get(p.get("section", "?"), 0) + 1
        if p.get("part_number_status") == "unconfirmed":
            ref = f"{p.get('section','?')}#{p.get('diagram_callout_id','?')}"
            unconfirmed_refs.append(ref)
    sec_names = [s["section_name"] for s in sections.get("sections", [])]
    prompt = (
        f"Round {round_no} for model '{mn}'. Target expected count: {expected}. "
        f"Master list now holds {len(master)} unique parts "
        f"(confirmed={by_status['confirmed']}, probable={by_status['probable']}, "
        f"unconfirmed={by_status['unconfirmed']}). "
        f"Per-section coverage so far: {json.dumps(coverage)}. "
        f"All sections: {json.dumps(sec_names)}. "
        f"Unconfirmed parts needing an MPN: {json.dumps(unconfirmed_refs[:60])}.\n\n"
        f"Past covered sections (DO NOT assign these again): {json.dumps(past_covered)}.\n\n"
        f"Produce assignments for {WORKERS_STEP6} workers (W1..W{WORKERS_STEP6}) for the next "
        "extraction round. Make the lanes DISJOINT — no two workers should target the same section "
        "or callout range — to minimize overlap. Prioritize under-covered sections and resolving "
        "unconfirmed MPNs (put those refs in resolve_unconfirmed for the relevant worker). "
        "Also return still_missing_estimate = expected count minus current unique count (floored at 0)."
    )
    return _call(MODEL_LITE, prompt, S7_GAMEPLAN, "MINIMAL")


# --------------------------------------------------------------------------- dedup
_STATUS_RANK = {"confirmed": 3, "probable": 2, "unconfirmed": 1}


def dedup_key(part: dict):
    """Primary key = normalized MPN; fallback = section+callout; last resort = section+name."""
    mpn = (part.get("manufacturer_part_number") or "")
    mpn = re.sub(r"[\s\-]", "", mpn).upper()
    status = part.get("part_number_status")
    if mpn and status in ("confirmed", "probable"):
        return ("mpn", mpn)
    sec = (part.get("section") or "").strip().lower()
    cid = (part.get("diagram_callout_id") or "").strip().lower()
    if sec and cid:
        return ("callout", sec, cid)
    name = (part.get("part_name") or "").strip().lower()
    return ("name", sec, name)


def merge_parts(master: dict, provenance: dict, new_parts: list) -> int:
    """Merge worker output into master keyed by dedup_key. Stronger status wins on collision.
    `provenance[key]` accumulates {count, sources[]} across all workers/rounds so the
    scaffold's discovery_source_count and per-sighting part_observations can be rebuilt.
    Returns count of genuinely new keys added."""
    added = 0
    for p in new_parts:
        k = dedup_key(p)
        prov = provenance.setdefault(k, {"count": 0, "sources": []})
        prov["count"] += 1
        prov["sources"].append({
            "source_url": p.get("source_url"),
            "diagram_url": p.get("diagram_url"),
            "status": p.get("part_number_status"),
            "manufacturer_part_number": p.get("manufacturer_part_number"),
        })
        if k not in master:
            master[k] = p
            added += 1
        else:
            cur = master[k]
            if _STATUS_RANK.get(p.get("part_number_status"), 0) > \
               _STATUS_RANK.get(cur.get("part_number_status"), 0):
                master[k] = p  # upgrade unconfirmed -> confirmed
    return added


def initial_assignments(sections: dict) -> list:
    """Round 1 has no step-7 output: split sections round-robin across the swarm."""
    names = [s["section_name"] for s in sections.get("sections", [])] or ["All"]
    buckets = [[] for _ in range(WORKERS_STEP6)]
    for i, name in enumerate(names):
        buckets[i % WORKERS_STEP6].append(name)
    return [
        {"worker_id": f"W{i+1}", "target_sections": b, "callout_ranges": [],
         "focus_note": "initial split", "resolve_unconfirmed": []}
        for i, b in enumerate(buckets)
        if b
    ]


# --------------------------------------------------------------------------- latency and payload
def to_scaffold_payload(results: dict, job_id: str) -> dict:
    """
    Transforms the pipeline output into the exact JSON schema expected by Neon DB.
    Matches the canonical_bom_parts, diagram_sections, and appliance_identities structures.
    """
    nameplate = results.get("nameplate", {})
    
    appliance_type = nameplate.get("appliance_type") or "unknown"
    appliance_type_lower = appliance_type.lower()
    
    valid_classes = {"washer", "dryer", "refrigerator", "stove", "range", "oven", "freezer", "unknown"}
    appliance_class = "unknown"
    for vc in valid_classes:
        if vc in appliance_type_lower:
            appliance_class = vc
            break

    conf_map = {"high": 1.0, "medium": 0.7, "low": 0.3}
    identity_confidence = conf_map.get(nameplate.get("extraction_confidence", "low"), 0.3)

    identity = {
        "model_number": nameplate.get("model_number"),
        "serial_number": nameplate.get("serial_number"),
        "input_brand": nameplate.get("brand"),
        "resolved_manufacturer": nameplate.get("brand"),
        "product_type": appliance_type,
        "appliance_class": appliance_class,
        "identity_confidence": identity_confidence,
        "input_source": "api"
    }

    diagram_sections = []
    for s in results.get("sections", []):
        urls = s.get("diagram_urls", [])
        diagram_sections.append({
            "source_section_name": s.get("section_name"),
            "normalized_section_name": s.get("section_name"),
            "diagram_image_url": urls[0] if urls else None,
            "observed_part_count": s.get("observed_part_count")
        })

    canonical_bom_parts = []
    status_to_conf = {"confirmed": 1.0, "probable": 0.7, "unconfirmed": 0.3}
    provenance = results.get("_provenance", {})

    for p in results.get("parts", []):
        prov_key = dedup_key(p)
        prov_key_str = "|".join(map(str, prov_key))
        prov_data = provenance.get(prov_key_str, {"count": 1, "sources": []})
        
        canonical_bom_parts.append({
            "model_number": nameplate.get("model_number"),
            "serial_number": nameplate.get("serial_number"),
            "manufacturer": nameplate.get("brand"),
            "appliance_class": appliance_class,
            "section_source_name": p.get("section"),
            "section_normalized": p.get("section"),
            "diagram_ref": p.get("diagram_callout_id"),
            "diagram_image_url": p.get("diagram_url"),
            "discovered_part_number": p.get("manufacturer_part_number"),
            "manufacturer_part_number": p.get("manufacturer_part_number"),
            "part_title": p.get("part_name"),
            "discovery_source_count": prov_data.get("count", 1),
            "part_identity_confidence": status_to_conf.get(p.get("part_number_status", "unconfirmed"), 0.3),
            "verification_status": "single_source" if prov_data.get("count", 1) == 1 else "cross_verified"
        })

    return {
        "job_id": job_id,
        "identity": identity,
        "diagram_sections": diagram_sections,
        "canonical_bom_parts": canonical_bom_parts,
        "expected_parts_count": results.get("expected_parts_count"),
        "expected_count_meta": results.get("expected_count_meta", {}),
        "parts_found": results.get("parts_found")
    }

def run_pipeline_fast(timeout_seconds: int = 55, audit_sections: bool = False,
                      **nameplate_input) -> dict:
    """
    Cold-sync path: runs the pipeline with a strict latency envelope.
    If it hits timeout_seconds, it will return partial results seamlessly.

    Bench logging policy:
        - LOCAL/DEV only: writes latency.log into services/extraction/cache/.
          Set PIPELINE_ENV=dev (or omit it) to enable.
        - PRODUCTION: set PIPELINE_ENV=production to suppress all file I/O.
          services/extraction/cache/ is NOT a durable store in serverless/Vercel;
          do not rely on it across invocations. Production cache → Neon extraction_cache.
    """
    _is_dev = os.environ.get("PIPELINE_ENV", "dev") != "production" and os.environ.get("VERCEL") != "1"

    bench_log = None
    if _is_dev:
        _cache_dir = os.path.join(os.path.dirname(__file__), "cache")
        os.makedirs(_cache_dir, exist_ok=True)
        bench_log = open(os.path.join(_cache_dir, "latency.log"), "a", encoding="utf-8")

    def log_bench(step, t0):
        t1 = time.perf_counter()
        if bench_log:
            bench_log.write(f"[{step}] {t1 - t0:.2f}s\n")
            bench_log.flush()
        return t1

    t_start = time.perf_counter()
    if bench_log:
        bench_log.write(f"\n--- Job Start {t_start} ---\n")

    print("[1] nameplate ingestion ...")
    nameplate = step1_nameplate(**nameplate_input)
    t_curr = log_bench("step1_nameplate", t_start)

    with ThreadPoolExecutor(max_workers=1) as count_executor:
        count_future = count_executor.submit(step5_expected_count, nameplate)

        print("[2] finding diagrams ...")
        diagrams = step2_find_diagrams(nameplate)
        t_curr = log_bench("step2_find_diagrams", t_curr)

        print("[3] consolidating sections ...")
        sections = (
            step3_consolidate(nameplate, diagrams)
            if audit_sections
            else consolidate_diagram_results(diagrams)
        )
        t_curr = log_bench("step3_consolidate", t_curr)

        if audit_sections:
            print("[4] auditing for missing sections ...")
            audits = step4_audit(nameplate, sections)
            sections = merge_audit_into_sections(sections, audits)
            t_curr = log_bench("step4_audit", t_curr)

        master: dict = {}
        provenance: dict = {}
        assignments = initial_assignments(sections)
        past_covered = set()
        consecutive_stalls = 0
        found_keys = set()

        print("[5/6] expected count and initial extraction ...")
        initial_jobs = [
            (lambda a=a: step6_extract(nameplate, sections, a, found_keys))
            for a in assignments
        ]
        initial_worker_outputs = _parallel(initial_jobs)
        raw_count_info = count_future.result()
    expected, count_info = normalize_expected_count(raw_count_info)
    if expected <= 0:
        raise RuntimeError(
            f"Expected part count is required before extracting {nameplate['model_number']}"
        )
    extraction_target = expected or min(
        MAX_CREDIBLE_EXPECTED_PARTS,
        max(PARTS_PER_WORKER, len(sections.get("sections", [])) * PARTS_PER_WORKER),
    )
    t_curr = log_bench("step5_count_and_initial_extract", t_curr)

    for rnd in range(1, MAX_ROUNDS + 1):
        if rnd == 1:
            worker_outputs = initial_worker_outputs
        else:
            elapsed = time.perf_counter() - t_start
            if elapsed > timeout_seconds - 15:
                # Leave 15s buffer for remaining logic/API responses.
                print("Timeout envelope approached. Returning partial.")
                if bench_log:
                    bench_log.write(
                        f"TIMEOUT EXCEEDED at round {rnd} (elapsed {elapsed:.2f}s)\n"
                    )
                break

            found_keys = {"|".join(map(str, k)) for k in master.keys()}
            print(f"[6] round {rnd}: extracting ...")
            jobs = [
                (lambda a=a: step6_extract(nameplate, sections, a, found_keys))
                for a in assignments
            ]
            worker_outputs = _parallel(jobs)
            t_curr = log_bench(f"step6_extract_r{rnd}", t_curr)

        added_total = 0
        success_count = 0
        for i, out in enumerate(worker_outputs):
            if isinstance(out, Exception) or "_error" in out:
                if bench_log:
                    err_msg = out if isinstance(out, Exception) else out.get("_error")
                    bench_log.write(f"Worker {i} failed: {err_msg}\n")
                continue
            
            success_count += 1
            # Mark successfully executed lanes as covered
            assigned_scope = assignments[i].get("target_sections", []) if i < len(assignments) else []
            for sec in assigned_scope:
                past_covered.add(sec)

            added_total += merge_parts(master, provenance, out.get("parts", []))
        
        if expected > 0 and len(master) >= expected:
            break

        # Saturation check: workers successfully ran, but nothing new was added
        if added_total == 0 and success_count > 0:
            consecutive_stalls += 1
            if consecutive_stalls >= STALL_ROUNDS:
                if bench_log:
                    bench_log.write(f"Saturation reached at round {rnd} (no new parts for {STALL_ROUNDS} rounds). Exiting.\n")
                break
        elif added_total > 0:
            consecutive_stalls = 0

        if rnd == MAX_ROUNDS:
            break

        print("[7] gameplan ...")
        plan = step7_gameplan(
            nameplate, sections, master, extraction_target, rnd + 1, list(past_covered)
        )
        assignments = plan.get("assignments") or initial_assignments(sections)
        
        # Code-side filter: drop overlapping sections
        filtered_assignments = []
        assigned_this_round = set()
        for a in assignments:
            valid_targets = []
            for sec in a.get("target_sections", []):
                # Only keep sections not covered previously, and not already assigned this round
                if sec not in past_covered and sec not in assigned_this_round:
                    valid_targets.append(sec)
                    assigned_this_round.add(sec)
            
            if valid_targets:
                a["target_sections"] = valid_targets
                filtered_assignments.append(a)
            elif a.get("resolve_unconfirmed"):
                # Even if target_sections are empty, if there's unconfirmed parts to resolve, keep it
                a["target_sections"] = []
                filtered_assignments.append(a)

        assignments = filtered_assignments
        if not assignments:
            if bench_log:
                bench_log.write(f"No new disjoint assignments available at round {rnd}. Exiting.\n")
            break

        t_curr = log_bench(f"step7_gameplan_r{rnd}", t_curr)

    parts = list(master.values())
    prov_json = {"|".join(map(str, k)): v for k, v in provenance.items()}
    
    total_time = time.perf_counter() - t_start
    if bench_log:
        bench_log.write(f"--- Job End (Total: {total_time:.2f}s) ---\n")
        bench_log.close()

    return {
        "nameplate": nameplate,
        "expected_parts_count": expected,
        "expected_count_meta": count_info,
        "sections": sections.get("sections", []),
        "parts_found": len(parts),
        "parts": parts,
        "_provenance": prov_json,
    }


# --------------------------------------------------------------------------- pipeline
def run_pipeline(**nameplate_input) -> dict:
    # Use the fast pipeline logic for both by default to ensure latency limits
    return run_pipeline_fast(timeout_seconds=9999, audit_sections=True, **nameplate_input)


def run_pipeline_warm(**nameplate_input) -> dict:
    """
    Warm-cache path: intended to read a pre-computed result from the Neon
    extraction_cache table before falling back to a live extraction run.

    Production cache (Neon extraction_cache) is a follow-up task.
    Until that table exists this is a transparent pass-through to
    run_pipeline_fast so the calling endpoint doesn't need to change later.

    When the Neon cache layer is wired:
        1. Look up model_number in extraction_cache.
        2. If cache hit and age < TTL: return cached payload directly.
        3. On miss: call run_pipeline_fast, store result, return it.
    """
    # TODO(follow-up): query Neon extraction_cache before running extraction.
    return run_pipeline_fast(audit_sections=True, **nameplate_input)


# ----------------------------------------------------------------- input + export
# Input CSV header aliases -> canonical field. Header matching is case-insensitive.
INPUT_ALIASES = {
    "model_number": ["model_number", "model", "model#", "model no", "modelnumber", "mn"],
    "serial": ["serial", "serial_number", "serial#", "sn"],
    "brand": ["brand", "make", "manufacturer", "mfg"],
    "appliance_type": ["appliance_type", "type", "category"],
}

# Per-part columns for the flat export (model_number is prepended).
PART_FIELDS = [
    "section", "diagram_callout_id", "manufacturer_part_number", "part_number_status",
    "part_name", "quantity", "is_sub_assembly", "source_url", "diagram_url", "notes",
]


def read_input_csv(path: str) -> list:
    """Read a CSV of appliances. Only model_number is required; serial/brand/type optional."""
    with open(path, newline="", encoding="utf-8-sig") as f:  # utf-8-sig strips Excel BOM
        reader = csv.DictReader(f)
        hmap = {}
        for col in reader.fieldnames or []:
            key = col.strip().lower()
            for canon, aliases in INPUT_ALIASES.items():
                if key in aliases:
                    hmap[col] = canon
        rows = []
        for raw in reader:
            row = {}
            for col, val in raw.items():
                if col in hmap and val and val.strip():
                    row[hmap[col]] = val.strip()
            if row.get("model_number"):
                rows.append(row)
    return rows


def _part_rows(results: list):
    for r in results:
        mn = r.get("nameplate", {}).get("model_number", "")
        for p in r.get("parts", []):
            row = {"model_number": mn}
            for fld in PART_FIELDS:
                row[fld] = p.get(fld)
            yield row


def export_csv(results: list, path: str) -> None:
    headers = ["model_number"] + PART_FIELDS
    n = 0
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for row in _part_rows(results):
            w.writerow(row)
            n += 1
    print("    wrote %s (%d rows)" % (path, n))


def export_xlsx(results: list, path: str) -> None:
    """Parts sheet (status colour-coded, frozen header, autofilter) + Summary sheet."""
    if openpyxl is None:
        print("    (openpyxl not installed; skipped .xlsx — `pip install openpyxl`)")
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"
    headers = ["model_number"] + PART_FIELDS
    ws.append(headers)

    fills = {
        "confirmed": PatternFill("solid", fgColor="C6EFCE"),   # green
        "probable": PatternFill("solid", fgColor="FFEB9C"),    # amber
        "unconfirmed": PatternFill("solid", fgColor="FFC7CE"),  # red
    }
    status_col = headers.index("part_number_status") + 1
    for row in _part_rows(results):
        ws.append([row.get(h) for h in headers])
        st = row.get("part_number_status")
        if st in fills:
            ws.cell(row=ws.max_row, column=status_col).fill = fills[st]

    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(headers[c - 1]) + 2, 12), 42)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), max(ws.max_row, 1))

    ws2 = wb.create_sheet("Summary")
    ws2.append(["model_number", "expected", "found", "confirmed", "probable", "unconfirmed"])
    for r in results:
        counts = {"confirmed": 0, "probable": 0, "unconfirmed": 0}
        for p in r.get("parts", []):
            s = p.get("part_number_status", "unconfirmed")
            counts[s] = counts.get(s, 0) + 1
        ws2.append([
            r.get("nameplate", {}).get("model_number", ""),
            r.get("expected_parts_count"), r.get("parts_found"),
            counts["confirmed"], counts["probable"], counts["unconfirmed"],
        ])
    for c in range(1, 7):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    wb.save(path)
    print("    wrote %s" % path)


def run_batch(input_csv: str, out_dir: str = "batch_output") -> list:
    """Run the full pipeline for every model_number in input_csv; write per-model JSON + combined exports."""
    rows = read_input_csv(input_csv)
    if not rows:
        print("No model_number rows found in", input_csv)
        return []
    os.makedirs(out_dir, exist_ok=True)
    results = []
    for i, row in enumerate(rows, 1):
        mn = row["model_number"]
        print("\n=== [%d/%d] %s ===" % (i, len(rows), mn))
        try:
            res = run_pipeline(
                model_number=mn, serial=row.get("serial"),
                brand=row.get("brand"), appliance_type=row.get("appliance_type"),
            )
        except Exception as e:  # one bad model must not kill the batch
            print("  pipeline failed:", e)
            res = {"nameplate": {"model_number": mn}, "expected_parts_count": 0,
                   "parts_found": 0, "parts": [], "error": str(e)}
            results.append(res)
            continue
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", mn)
        with open(os.path.join(out_dir, safe + ".json"), "w", encoding="utf-8") as f:
            json.dump(res, f, indent=2, ensure_ascii=False)
        results.append(res)

    export_csv(results, os.path.join(out_dir, "all_parts.csv"))
    export_xlsx(results, os.path.join(out_dir, "all_parts.xlsx"))
    print("\nBatch done: %d models, %d total parts -> %s/" %
          (len(results), sum(r["parts_found"] for r in results), out_dir))
    return results


# ------------------------------------------------------------------------------ entry
if __name__ == "__main__":
    if len(sys.argv) > 1:
        # Batch:  python pipeline.py models.csv [out_dir]
        run_batch(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "batch_output")
    else:
        # Single example (manual entry). Photo / OCR options shown commented.
        result = run_pipeline(
            model_number="WTW5000DW1",
            brand="Whirlpool",
            appliance_type="Top-load Washer",
        )
        # result = run_pipeline(image_path="nameplate.jpg")
        # result = run_pipeline(raw_text="MODEL WTW5000DW1  SER C71234567  WHIRLPOOL")

        with open("parts_list.json", "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        export_csv([result], "parts_list.csv")
        export_xlsx([result], "parts_list.xlsx")
        print("\nWrote %d parts (json / csv / xlsx)" % result["parts_found"])
